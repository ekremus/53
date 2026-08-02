#!/usr/bin/env python3
"""Build audited before/after snapshots from the replay workbook.

This script is read-only with respect to production. It fetches the current
public state, converts the workbook rows, and writes migration artifacts.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import unicodedata
import urllib.request
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


MONTHS = {
    "Ocak": 1,
    "Şubat": 2,
    "Mart": 3,
    "Nisan": 4,
    "Mayıs": 5,
    "Haziran": 6,
    "Temmuz": 7,
    "Ağustos": 8,
    "Eylül": 9,
    "Ekim": 10,
    "Kasım": 11,
    "Aralık": 12,
}

CIVILIZATIONS = {
    "Birmanyalılar": "Burmese",
    "Bizanslılar": "Byzantines",
    "Bohemyalılar": "Bohemians",
    "Britanyalılar": "Britons",
    "Bulgarlar": "Bulgarians",
    "Dravidler": "Dravidians",
    "Franklar": "Franks",
    "Gotlar": "Goths",
    "Gurjaralar": "Gurjaras",
    "Hindustanlılar": "Hindustanis",
    "Hunlar": "Huns",
    "İnkalar": "Incas",
    "İspanyollar": "Spanish",
    "İtalyanlar": "Italians",
    "Keltler": "Celts",
    "Khmerler": "Khmer",
    "Koreliler": "Koreans",
    "Kumanlar": "Cumans",
    "Litvanyalılar": "Lithuanians",
    "Macarlar": "Magyars",
    "Makedonlar": "Macedonians",
    "Malaylar": "Malay",
    "Moğollar": "Mongols",
    "Persler": "Persians",
    "Polonyalılar": "Poles",
    "Puru": "Puru",
    "Romalılar": "Romans",
    "Sarazenler": "Saracens",
    "Slavlar": "Slavs",
    "Tatarlar": "Tatars",
    "Trakyalılar": "Thracians",
    "Tötonlar": "Teutons",
    "Türkler": "Turks",
    "Çinliler": "Chinese",
}


def normalized_name(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.strip().replace("ı", "i"))
    return "".join(character for character in decomposed if not unicodedata.combining(character)).casefold()


def parse_timestamp(value: str) -> datetime:
    date_part, time_part = (part.strip() for part in value.split("·"))
    day, month_name, year = date_part.split()
    hour, minute = (int(part) for part in time_part.split(":"))
    return datetime(int(year), MONTHS[month_name], int(day), hour, minute)


def match_date(timestamp: datetime) -> str:
    session_timestamp = timestamp - timedelta(days=1) if timestamp.hour < 6 else timestamp
    return session_timestamp.date().isoformat()


def split_slot(value: str) -> tuple[str, str]:
    player, civilization = (part.strip() for part in value.rsplit(" — ", 1))
    if civilization not in CIVILIZATIONS:
        raise ValueError(f"Bilinmeyen uygarlık: {civilization}")
    return player, CIVILIZATIONS[civilization]


def state_digest(state: dict) -> str:
    serialized = json.dumps(state, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def semantic_match(match: dict) -> dict:
    return {
        "date": match["date"],
        "winner": match["winner"],
        "teams": {
            team_id: sorted(
                match["teams"][team_id],
                key=lambda slot: (slot["playerId"], slot["civilization"]),
            )
            for team_id in ("cortinyanlar", "bakracogullari")
        },
    }


def parse_workbook(path: Path, players_by_name: dict[str, str]) -> list[dict]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    if worksheet.max_column != 4 or worksheet.cell(1, 1).value != "#":
        raise ValueError("Replay çalışma sayfası beklenen dört sütunlu yapıda değil.")

    records = []
    for row in range(2, worksheet.max_row + 1, 4):
        source_number = worksheet.cell(row, 1).value
        timestamp = parse_timestamp(worksheet.cell(row, 2).value)
        winning_slots = [split_slot(worksheet.cell(slot_row, 3).value) for slot_row in range(row, row + 4)]
        losing_slots = [split_slot(worksheet.cell(slot_row, 4).value) for slot_row in range(row, row + 4)]

        winner_has_alman = any(player == "Alman General" for player, _ in winning_slots)
        loser_has_alman = any(player == "Alman General" for player, _ in losing_slots)
        winner_has_mstf = any(player == "mstfunsal" for player, _ in winning_slots)
        loser_has_mstf = any(player == "mstfunsal" for player, _ in losing_slots)
        if winner_has_alman == loser_has_alman or winner_has_mstf == loser_has_mstf:
            raise ValueError(f"{source_number}. maçta takım kaptanları ayırt edilemiyor.")
        if winner_has_alman == winner_has_mstf:
            raise ValueError(f"{source_number}. maçta takım kaptanları aynı tarafta.")

        winner = "cortinyanlar" if winner_has_alman else "bakracogullari"
        cortinyanlar = winning_slots if winner == "cortinyanlar" else losing_slots
        bakracogullari = winning_slots if winner == "bakracogullari" else losing_slots

        def slots(values: list[tuple[str, str]]) -> list[dict]:
            converted = []
            for player_name, civilization in values:
                player_id = players_by_name.get(normalized_name(player_name))
                if not player_id:
                    raise ValueError(f"Kayıtlı olmayan oyuncu: {player_name}")
                converted.append({"playerId": player_id, "civilization": civilization})
            return converted

        records.append({
            "sourceNumber": source_number,
            "timestamp": timestamp,
            "match": {
                "id": f"replay-{timestamp:%Y%m%d-%H%M}-{source_number:03d}",
                "date": match_date(timestamp),
                "teams": {
                    "cortinyanlar": slots(cortinyanlar),
                    "bakracogullari": slots(bakracogullari),
                },
                "winner": winner,
            },
        })
    return records


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=Path("replays/aoe2_replays.xlsx"))
    parser.add_argument("--state-url", default="https://53aoe.vercel.app/api/state")
    parser.add_argument("--output", type=Path, default=Path("data/migrations/2026-08-02-replays"))
    args = parser.parse_args()

    request = urllib.request.Request(args.state_url, headers={"Cache-Control": "no-store"})
    with urllib.request.urlopen(request) as response:
        before = json.load(response)["state"]

    players_by_name = {normalized_name(player["name"]): player["id"] for player in before["players"]}
    records = parse_workbook(args.workbook, players_by_name)
    existing_matches = {json.dumps(semantic_match(match), ensure_ascii=False, sort_keys=True) for match in before["matches"]}
    duplicate_rows = []
    imported_records = []
    for record in records:
        signature = json.dumps(semantic_match(record["match"]), ensure_ascii=False, sort_keys=True)
        if signature in existing_matches:
            duplicate_rows.append(record["sourceNumber"])
        else:
            imported_records.append(record)

    # The app resolves same-day display order from array position. Append older
    # replay rows first so the latest replay in a session receives the later index.
    after = copy.deepcopy(before)
    after["matches"].extend(record["match"] for record in reversed(imported_records))

    gaps = [
        int((newer["timestamp"] - older["timestamp"]).total_seconds() // 60)
        for newer, older in zip(records, records[1:])
    ]
    source_bytes = args.workbook.read_bytes()
    report = {
        "source": str(args.workbook),
        "sourceSha256": hashlib.sha256(source_bytes).hexdigest(),
        "sourceMatches": len(records),
        "duplicateSourceRows": duplicate_rows,
        "importedMatches": len(imported_records),
        "importedSourceRows": [record["sourceNumber"] for record in imported_records],
        "importedDateRange": {
            "from": min(record["match"]["date"] for record in imported_records),
            "to": max(record["match"]["date"] for record in imported_records),
        },
        "importedWinnerTotals": dict(Counter(record["match"]["winner"] for record in imported_records)),
        "durationColumnPresent": False,
        "underTenMinuteMatchesDetectedFromTimestamps": sum(gap < 10 for gap in gaps),
        "minimumAdjacentReplayGapMinutes": min(gaps),
        "excludedUnderTenMinuteSourceRows": [],
        "before": {
            "digest": state_digest(before),
            "revision": before["revision"],
            "players": len(before["players"]),
            "matches": len(before["matches"]),
        },
        "after": {
            "digest": state_digest(after),
            "revisionBeforeServerWrite": after["revision"],
            "players": len(after["players"]),
            "matches": len(after["matches"]),
        },
    }

    write_json(args.output / "before.json", before)
    write_json(args.output / "after.json", after)
    write_json(args.output / "report.json", report)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
